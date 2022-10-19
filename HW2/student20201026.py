#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

list_student = {}
row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		list_student[row_id] = sum_v
	row_id += 1

sort_list_student = sorted(list_student.items(), key=lambda item: item[1], reverse=True)
sort_list_index = []
sort_list_values = []
for i in sort_list_student:
	sort_list_index.append(i[0])
	sort_list_values.append(i[1])
sorted_list_student = dict(zip(sort_list_index, sort_list_values))

count = len(list_student)+1
for i in range(len(list_student)):
	if count*0.3*0.5-1 >= i:
		sorted_list_student[sort_list_index[i]] = 'A+'
	elif count*0.3-1 >= i:
		sorted_list_student[sort_list_index[i]] = 'A0'
	elif count*0.4*0.5-1+count*0.3 >= i:
		sorted_list_student[sort_list_index[i]] = 'B+'
	elif count*0.7-1 >= i:
		sorted_list_student[sort_list_index[i]] = 'B0'
	else:
		if count*0.3*0.5-1+count*0.7 >= i:
			sorted_list_student[sort_list_index[i]] = 'C+'
		else:	
			sorted_list_student[sort_list_index[i]] = 'C0'

sorted_list_grade = []
for value in sorted_list_student.values():
	sorted_list_grade.append(value)

row_id = 1 
for row in ws:
	if row_id != 1:
		ws.cell(row = sort_list_index[row_id-2], column = 8, value = sorted_list_grade[row_id-2])
	row_id += 1

#grade
	
wb.save( "student.xlsx" )

